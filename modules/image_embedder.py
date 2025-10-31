# modules/image_embedder.py
import pandas as pd
import requests
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from io import BytesIO
from PIL import Image, ImageFile
import base64
from datetime import datetime
import os
import re
from typing import Dict, List, Optional, Callable, Any
from pathlib import Path
import logging
import time

from .path_utils import get_logs_path, get_cache_path

# Allow loading truncated images
ImageFile.LOAD_TRUNCATED_IMAGES = True


class ImageEmbedder:
    """
    Image embedding functionality for Excel files.
    Converts image URLs and base64 data to embedded images in Excel cells.
    """
    
    def __init__(self):
        self.is_running = False
        self.config = {}
        self.callbacks = {}
        self.session = None
        self.image_cache = {}
        self.processed_count = 0
        self.error_count = 0
        self.df = None
        self.image_columns = []
        self.base64_columns = []
        
        # Setup module-specific logger
        self.logger = self._setup_logger()
    
    def _setup_logger(self) -> logging.Logger:
        """Setup module-specific logger."""
        logger = logging.getLogger(__name__)
        if not logger.handlers:
            log_file = get_logs_path() / "image_embedder.log"
            handler = logging.FileHandler(log_file, encoding='utf-8')
            formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            )
            handler.setFormatter(formatter)
            logger.addHandler(handler)
            logger.setLevel(logging.INFO)
        return logger
    
    def configure(self,
                  input_file: str,
                  output_file: str,
                  max_width: int = 100,
                  max_height: int = 100,
                  image_quality: int = 85,
                  row_height: int = 75,
                  column_width: int = 15,
                  timeout: int = 30,
                  enable_cache: bool = True,
                  log_callback: Callable = None,
                  status_callback: Callable = None,
                  progress_callback: Callable = None,
                  stop_callback: Callable = None) -> None:
        """
        Configure the image embedder module.
        
        Args:
            input_file: Path to input Excel file
            output_file: Path to output Excel file
            max_width: Maximum image width in pixels
            max_height: Maximum image height in pixels
            image_quality: JPEG quality (1-100)
            row_height: Excel row height
            column_width: Excel column width
            timeout: Request timeout in seconds
            enable_cache: Enable image caching
            log_callback: Callback for log messages
            status_callback: Callback for status updates
            progress_callback: Callback for progress updates
            stop_callback: Callback to check if processing should stop
        """
        self.config = {
            'input_file': input_file,
            'output_file': output_file,
            'max_width': max_width,
            'max_height': max_height,
            'image_quality': image_quality,
            'row_height': row_height,
            'column_width': column_width,
            'timeout': timeout,
            'enable_cache': enable_cache
        }
        
        self.callbacks = {
            'log': log_callback,
            'status': status_callback,
            'progress': progress_callback,
            'stop': stop_callback
        }
        
        # Initialize requests session for connection pooling
        if enable_cache:
            self.session = requests.Session()
            self.session.headers.update({
                'User-Agent': 'PhoneLookupTool/1.0.0'
            })
        
        self.logger.info(f"ImageEmbedder configured: input={input_file}, output={output_file}")
        self.log("Image embedder module configured successfully")
    
    def should_stop(self) -> bool:
        """Check if processing should stop."""
        if self.callbacks.get('stop'):
            return self.callbacks['stop']()
        return False
    
    def log(self, message: str, level: str = "info", also_print: bool = False) -> None:
        """Log message to both file and callback."""
        # Log to file
        if level == "info":
            self.logger.info(message)
        elif level == "error":
            self.logger.error(message)
        elif level == "warning":
            self.logger.warning(message)
        
        # Send to callback
        if self.callbacks.get('log'):
            self.callbacks['log'](message)
        
        # Handle the also_print parameter that tests are using
        if also_print:
            print(f"[{level.upper()}] {message}")
    
    def update_status(self, message: str) -> None:
        """Update status message."""
        if self.callbacks.get('status'):
            self.callbacks['status'](message)
    
    def update_progress(self, value: float) -> None:
        """Update progress percentage."""
        if self.callbacks.get('progress'):
            self.callbacks['progress'](value)
    
    def run(self) -> bool:
        """
        Run the image embedding process.
        
        Returns:
            bool: True if successful, False otherwise
        """
        self.is_running = True
        self.processed_count = 0
        self.error_count = 0
        
        try:
            self.log("Starting image embedding process")
            success = self._embed_images()
            
            if success:
                self.log(f"Image embedding completed successfully. Processed: {self.processed_count}, Errors: {self.error_count}")
            else:
                self.log(f"Image embedding completed with issues. Processed: {self.processed_count}, Errors: {self.error_count}")
            
            return success
            
        except Exception as e:
            error_msg = f"Unexpected error in image embedding: {str(e)}"
            self.log(error_msg, "error")
            self.update_status("Image embedding failed with unexpected error")
            return False
            
        finally:
            self.is_running = False
            # Clear cache and close session
            self.image_cache.clear()
            if self.session:
                self.session.close()
    
    def stop(self) -> None:
        """Stop the image embedding process."""
        self.is_running = False
        self.log("Stop signal received")
    
    def _embed_images(self) -> bool:
        """Perform the actual image embedding process."""
        try:
            # Step 1: Load and validate input file
            self.update_status("Loading Excel file...")
            df, image_cols, b64_cols = self._load_input_file()
            if df is None:
                return False
            
            self.df = df
            self.image_columns = image_cols
            self.base64_columns = b64_cols
            
            total_images = self._count_total_images()
            self.log(f"Found {len(image_cols)} image columns and {len(b64_cols)} base64 columns")
            self.log(f"Total images to process: {total_images}")
            
            # Step 2: Create or load workbook
            self.update_status("Preparing workbook...")
            wb, ws = self._prepare_workbook()
            if wb is None:
                return False
            
            # Step 3: Configure Excel layout
            self._configure_excel_layout(ws, image_cols, b64_cols)
            
            # Step 4: Process images row by row
            self.update_status("Embedding images...")
            processed_images = 0
            
            for row_idx, row in df.iterrows():
                if self.should_stop():
                    self.log("Image embedding stopped by user")
                    # Save the workbook before returning when stopped gracefully
                    try:
                        self._save_workbook(wb)
                    except:
                        pass
                    return True  # Return True when stopped gracefully
                
                excel_row_idx = row_idx + 2  # +1 for header, +1 for 1-based indexing
                processed_in_row = self._process_row_images(ws, excel_row_idx, row, image_cols, b64_cols)
                processed_images += processed_in_row
                
                # Update progress
                if total_images > 0:
                    progress = (processed_images / total_images) * 100
                    self.update_progress(progress)
                    self.update_status(f"Embedding images... {processed_images}/{total_images}")
                
                # Log progress every 10 rows
                if (row_idx + 1) % 10 == 0:
                    self.log(f"Processed {row_idx + 1} rows, {processed_images} images")
            
            # Step 5: Save workbook
            if not self.should_stop():
                self.update_status("Saving workbook...")
                success = self._save_workbook(wb)
                if success:
                    self.update_status("Image embedding completed successfully")
                    return True
                else:
                    self.update_status("Failed to save workbook")
                    return False
            else:
                self.update_status("Image embedding stopped")
                return True  # Return True when stopped gracefully
                
        except Exception as e:
            error_msg = f"Error during image embedding: {str(e)}"
            self.log(error_msg, "error")
            import traceback
            self.log(f"Traceback: {traceback.format_exc()}", "error")
            return False

    def _load_input_file(self) -> tuple:
        """
        Load and validate the input Excel file.
        
        Returns:
            tuple: (dataframe, image_columns, b64_columns) or (None, [], [])
        """
        try:
            df = pd.read_excel(self.config['input_file'])
            
            # Find image and base64 columns
            image_cols = [col for col in df.columns if col.startswith('Image_') or col.startswith('image_')]
            b64_cols = [col for col in df.columns if col.startswith('b64_') or col.startswith('base64_')]
            
            if not image_cols and not b64_cols:
                self.log("No image or base64 columns found in input file", "warning")
            
            self.log(f"Input file loaded: {len(df)} rows, {len(image_cols)} image columns, {len(b64_cols)} base64 columns")
            return df, image_cols, b64_cols
            
        except Exception as e:
            error_msg = f"Error loading input file: {str(e)}"
            self.log(error_msg, "error")
            self.update_status("Error loading input file")
            return None, [], []
    
    def _count_total_images(self, df: pd.DataFrame = None, image_cols: List[str] = None, b64_cols: List[str] = None) -> int:
        """Count total number of images to process."""
        # Use instance variables if parameters not provided
        if df is None:
            df = self.df
        if image_cols is None:
            image_cols = self.image_columns
        if b64_cols is None:
            b64_cols = self.base64_columns
        
        if df is None:
            return 0
            
        total = 0
        
        # Count image URLs - only count non-empty strings
        for col in image_cols:
            if col in df.columns:
                # Count only non-empty, non-whitespace strings
                total += int(df[col].apply(lambda x: pd.notna(x) and str(x).strip() != '').sum())
        
        # Count base64 images - only count non-empty strings
        for col in b64_cols:
            if col in df.columns:
                # Count only non-empty, non-whitespace strings
                total += int(df[col].apply(lambda x: pd.notna(x) and str(x).strip() != '').sum())
        
        return total    
    
    def _prepare_workbook(self):
        """Prepare or create the Excel workbook."""
        try:
            output_file = self.config['output_file']
            
            if os.path.exists(output_file):
                # Load existing workbook
                wb = openpyxl.load_workbook(output_file)
                self.log(f"Loaded existing workbook: {output_file}")
            else:
                # Create new workbook by copying the input file structure if available
                input_file = self.config.get('input_file')
                
                if input_file and os.path.exists(input_file):
                    # Copy the input file to output file to preserve all data
                    import shutil
                    shutil.copy2(input_file, output_file)
                    wb = openpyxl.load_workbook(output_file)
                    self.log(f"Copied input file to create workbook: {output_file}")
                else:
                    # Fallback: create empty workbook
                    wb = openpyxl.Workbook()
                    self.log(f"Created new workbook: {output_file}")
            
            ws = wb.active
            return wb, ws
            
        except Exception as e:
            error_msg = f"Error preparing workbook: {str(e)}"
            self.log(error_msg, "error")
            return None, None


    def _configure_excel_layout(self, ws, image_cols: List[str], b64_cols: List[str]) -> None:
        """Configure Excel row heights and column widths."""
        try:
            # Set row height for all rows that will contain images
            for row_idx in range(2, ws.max_row + 2):  # +2 for header and 1-based indexing
                ws.row_dimensions[row_idx].height = self.config['row_height']
            
            # Set column widths for image columns
            all_image_cols = image_cols + b64_cols
            for col_name in all_image_cols:
                try:
                    col_idx = self._get_column_index(ws, col_name)
                    if col_idx:
                        ws.column_dimensions[get_column_letter(col_idx)].width = self.config['column_width']
                except Exception:
                    # Column not found, skip
                    continue
            
            self.log("Excel layout configured successfully")
            
        except Exception as e:
            self.log(f"Error configuring Excel layout: {str(e)}", "warning")
    
    def _process_row_images(self, ws, excel_row_idx: int, row: pd.Series, 
                          image_cols: List[str], b64_cols: List[str]) -> int:
        """
        Process images for a single row.
        
        Returns:
            int: Number of images processed in this row
        """
        processed_count = 0
        
        # Process URL images
        for col_name in image_cols:
            if self.should_stop():
                return processed_count
            
            image_url = row[col_name]
            if pd.notna(image_url) and str(image_url).strip():
                col_idx = self._get_column_index(ws, col_name)
                if col_idx and self._embed_url_image(ws, excel_row_idx, col_idx, str(image_url).strip()):
                    processed_count += 1
        
        # Process base64 images
        for col_name in b64_cols:
            if self.should_stop():
                return processed_count
            
            b64_data = row[col_name]
            if pd.notna(b64_data) and str(b64_data).strip():
                col_idx = self._get_column_index(ws, col_name)
                if col_idx and self._embed_base64_image(ws, excel_row_idx, col_idx, str(b64_data).strip()):
                    processed_count += 1
        
        return processed_count
    
    def _get_column_index(self, ws, col_name: str) -> Optional[int]:
        """Get column index for a given column name."""
        try:
            header_row = ws[1]
            for idx, cell in enumerate(header_row, 1):
                if cell.value == col_name:
                    return idx
            return None
        except Exception:
            return None
    
    def _embed_url_image(self, ws, row_idx: int, col_idx: int, image_url: str) -> bool:
        """Embed an image from URL into Excel cell."""
        try:
            # Download image
            image_data = self._download_image(image_url)
            if not image_data:
                return False
            
            # Resize and embed
            success = self._embed_image_data(ws, row_idx, col_idx, image_data)
            if success:
                self.processed_count += 1
                self.log(f"Embedded URL image: {image_url}", also_print=False)
                return True
            else:
                self.error_count += 1
                return False
            
        except Exception as e:
            self.log(f"Error embedding URL image {image_url}: {str(e)}", "error")
            self.error_count += 1
            return False
    
    def _embed_base64_image(self, ws, row_idx: int, col_idx: int, b64_data: str) -> bool:
        """Embed a base64 image into Excel cell."""
        try:
            # Decode base64
            image_data = self._decode_base64_image(b64_data)
            if not image_data:
                return False
            
            # Resize and embed
            success = self._embed_image_data(ws, row_idx, col_idx, image_data)
            if success:
                self.processed_count += 1
                self.log(f"Embedded base64 image in row {row_idx}, column {col_idx}", also_print=False)
                return True
            else:
                self.error_count += 1
                return False
            
        except Exception as e:
            self.log(f"Error embedding base64 image: {str(e)}", "error")
            self.error_count += 1
            return False
    
    def _download_image(self, url: str) -> Optional[BytesIO]:
        """Download image from URL with caching."""
        # Check cache first
        if self.config['enable_cache'] and url in self.image_cache:
            self.log(f"Using cached image: {url}", also_print=False)
            return BytesIO(self.image_cache[url])
        
        try:
            if self.session:
                response = self.session.get(url, timeout=self.config['timeout'])
            else:
                response = requests.get(url, timeout=self.config['timeout'])
            
            response.raise_for_status()
            
            image_data = response.content
            
            # Cache the image
            if self.config['enable_cache']:
                self.image_cache[url] = image_data
            
            return BytesIO(image_data)
            
        except Exception as e:
            self.log(f"Error downloading image from {url}: {str(e)}", also_print=False)
            return None
    
    def _decode_base64_image(self, b64_string: str) -> Optional[BytesIO]:
        """
        Decode base64 image safely, automatically fixing padding or prefix issues.
        """
        try:
            if not b64_string or pd.isna(b64_string):
                return None

            # Remove potential data URL prefix
            b64_string = re.sub(r"^data:image\/[a-zA-Z]+;base64,", "", b64_string.strip())

            # Fix missing padding automatically
            missing_padding = len(b64_string) % 4
            if missing_padding:
                b64_string += "=" * (4 - missing_padding)

            decoded = base64.b64decode(b64_string, validate=False)
            return BytesIO(decoded)
            
        except Exception as e:
            self.log(f"Error decoding base64 image: {e}", also_print=False)
            return None
    
    def _embed_image_data(self, ws, row_idx: int, col_idx: int, image_data: BytesIO) -> bool:
        """Embed image data into Excel cell after resizing."""
        try:
            # Resize image
            resized_image = self._resize_image(image_data)
            if resized_image is None:
                return False
            
            # Create OpenPyXL image
            img = XLImage(resized_image)
            
            # Calculate cell coordinates
            cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
            
            # Add image to worksheet
            ws.add_image(img, cell_ref)
            
            return True
            
        except Exception as e:
            self.log(f"Error embedding image data: {str(e)}", also_print=False)
            return False
    
    def _resize_image(self, image_data: BytesIO) -> Optional[BytesIO]:
        """
        Resize image to fit Excel cell while maintaining aspect ratio.
        Converts image to RGB to avoid mode errors.
        """
        try:
            with Image.open(image_data) as img:
                # Convert to RGB if needed (avoids 'cannot write mode P as JPEG')
                if img.mode in ("P", "RGBA", "LA", "CMYK"):
                    img = img.convert("RGB")

                # Calculate new dimensions maintaining aspect ratio
                img.thumbnail((self.config['max_width'], self.config['max_height']))
                
                # Save as JPEG
                out_data = BytesIO()
                img.save(out_data, format="JPEG", quality=self.config['image_quality'])
                out_data.seek(0)
                return out_data
                
        except Exception as e:
            self.log(f"Error resizing image: {e}", also_print=False)
            return None
    
    def _save_workbook(self, wb) -> bool:
        """Save the workbook to file."""
        try:
            output_file = self.config['output_file']
            wb.save(output_file)
            self.log(f"Workbook saved successfully: {output_file}")
            return True
            
        except PermissionError:
            error_msg = f"Permission denied: Cannot save to {self.config['output_file']}. File may be open in Excel."
            self.log(error_msg, "error")
            self.update_status("Error: Close Excel file and try again")
            return False
            
        except Exception as e:
            error_msg = f"Error saving workbook: {str(e)}"
            self.log(error_msg, "error")
            self.update_status("Error saving workbook")
            return False
    
    def get_processing_stats(self) -> Dict[str, Any]:
        """Get current processing statistics."""
        return {
            "processed_count": self.processed_count,
            "error_count": self.error_count,
            "is_running": self.is_running
        }
    
    def __str__(self) -> str:
        """String representation."""
        stats = self.get_processing_stats()
        return f"ImageEmbedder(processed={stats['processed_count']}, errors={stats['error_count']}, running={stats['is_running']})"


# Utility functions
def embed_images_in_excel(input_file: str, 
                         output_file: str = None,
                         max_width: int = 100,
                         max_height: int = 100) -> bool:
    """
    Convenience function for quick image embedding.
    
    Args:
        input_file: Path to input Excel file
        output_file: Path to output Excel file (optional)
        max_width: Maximum image width
        max_height: Maximum image height
        
    Returns:
        bool: True if successful
    """
    if output_file is None:
        input_path = Path(input_file)
        output_file = input_path.parent / f"{input_path.stem}_with_images{input_path.suffix}"
    
    embedder = ImageEmbedder()
    embedder.configure(
        input_file=input_file,
        output_file=output_file,
        max_width=max_width,
        max_height=max_height
    )
    
    return embedder.run()


def validate_image_url(url: str) -> bool:
    """
    Validate if URL points to a likely image file.
    
    Args:
        url: URL to validate
        
    Returns:
        bool: True if URL appears to be an image
    """
    if not url or pd.isna(url):
        return False
    
    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']
    url_lower = url.lower()
    
    return any(ext in url_lower for ext in image_extensions)


if __name__ == "__main__":
    # Test the image embedder module
    print("Testing ImageEmbedder module...")
    
    # Create a test instance
    embedder = ImageEmbedder()
    
    # Test configuration
    embedder.configure(
        input_file="test_input.xlsx",
        output_file="test_output.xlsx",
        max_width=100,
        max_height=100
    )
    
    print(f"ImageEmbedder configured: {embedder}")
    print("Module test completed!")