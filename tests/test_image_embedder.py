# tests/test_image_embedder.py
import unittest
import pandas as pd
import tempfile
import os
from pathlib import Path
import sys
from unittest.mock import Mock, patch, MagicMock
import base64
from io import BytesIO
from PIL import Image
import openpyxl

# Add the parent directory to the path to import modules
sys.path.append(str(Path(__file__).parent.parent))

from modules.image_embedder import ImageEmbedder, embed_images_in_excel, validate_image_url


class TestImageEmbedder(unittest.TestCase):
    """Test cases for ImageEmbedder class."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.image_embedder = ImageEmbedder()
        self.test_input_file = Path(tempfile.mktemp(suffix=".xlsx"))
        self.test_output_file = Path(tempfile.mktemp(suffix=".xlsx"))
        
        # Create test data directory
        self.test_data_dir = Path(__file__).parent / "test_data"
        self.test_data_dir.mkdir(exist_ok=True)
        
        # Create test images
        self.test_image_url = "http://example.com/test.jpg"
        self.test_base64_image = self._create_test_base64_image()
        
        # Create test Excel file
        self._create_test_excel_file()
    
    def tearDown(self):
        """Clean up after tests."""
        # Remove test files if they exist
        for file_path in [self.test_input_file, self.test_output_file]:
            if file_path.exists():
                file_path.unlink()
    
    def _create_test_base64_image(self) -> str:
        """Create a test base64 encoded image."""
        # Create a simple test image
        img = Image.new('RGB', (100, 100), color='red')
        img_bytes = BytesIO()
        img.save(img_bytes, format='JPEG')
        img_bytes.seek(0)
        
        # Encode to base64
        encoded = base64.b64encode(img_bytes.getvalue()).decode('utf-8')
        return encoded
    
    def _create_test_excel_file(self):
        """Create a test Excel file with image URLs and base64 data."""
        test_data = {
            'Number': ['923001234567', '923001234568', '923001234569'],
            'Name': ['User One', 'User Two', 'User Three'],
            'Image_1': [self.test_image_url, '', 'http://example.com/another.jpg'],
            'Image_2': ['http://example.com/image2.jpg', '', ''],
            'b64_1': [self.test_base64_image, '', ''],
            'b64_2': ['', 'data:image/jpeg;base64,' + self.test_base64_image, ''],
            'Status': ['Success', 'Success', 'Success']
        }
        df = pd.DataFrame(test_data)
        df.to_excel(self.test_input_file, index=False)
    
    def test_initialization(self):
        """Test that ImageEmbedder initializes correctly."""
        self.assertIsInstance(self.image_embedder, ImageEmbedder)
        self.assertFalse(self.image_embedder.is_running)
        self.assertEqual(self.image_embedder.processed_count, 0)
        self.assertEqual(self.image_embedder.error_count, 0)
        self.assertEqual(len(self.image_embedder.image_cache), 0)
    
    def test_configure(self):
        """Test configuration of ImageEmbedder."""
        config = {
            'input_file': str(self.test_input_file),
            'output_file': str(self.test_output_file),
            'max_width': 150,
            'max_height': 150,
            'image_quality': 90,
            'row_height': 80,
            'column_width': 20,
            'timeout': 15,
            'enable_cache': False
        }
        
        self.image_embedder.configure(
            input_file=config['input_file'],
            output_file=config['output_file'],
            max_width=config['max_width'],
            max_height=config['max_height'],
            image_quality=config['image_quality'],
            row_height=config['row_height'],
            column_width=config['column_width'],
            timeout=config['timeout'],
            enable_cache=config['enable_cache']
        )
        
        # Check configuration
        self.assertEqual(self.image_embedder.config['input_file'], config['input_file'])
        self.assertEqual(self.image_embedder.config['output_file'], config['output_file'])
        self.assertEqual(self.image_embedder.config['max_width'], config['max_width'])
        self.assertEqual(self.image_embedder.config['max_height'], config['max_height'])
        self.assertEqual(self.image_embedder.config['image_quality'], config['image_quality'])
        self.assertEqual(self.image_embedder.config['enable_cache'], config['enable_cache'])
    
    def test_configure_with_callbacks(self):
        """Test configuration with callbacks."""
        mock_log = Mock()
        mock_status = Mock()
        mock_progress = Mock()
        mock_stop = Mock(return_value=False)
        
        self.image_embedder.configure(
            input_file=str(self.test_input_file),
            output_file=str(self.test_output_file),
            max_width=100,
            max_height=100,
            log_callback=mock_log,
            status_callback=mock_status,
            progress_callback=mock_progress,
            stop_callback=mock_stop
        )
        
        # Test callbacks are set
        self.assertEqual(self.image_embedder.callbacks['log'], mock_log)
        self.assertEqual(self.image_embedder.callbacks['status'], mock_status)
        self.assertEqual(self.image_embedder.callbacks['progress'], mock_progress)
        self.assertEqual(self.image_embedder.callbacks['stop'], mock_stop)
        
        # Test stop callback
        self.assertFalse(self.image_embedder.should_stop())
        mock_stop.assert_called_once()
    
    def test_load_input_file(self):
        """Test loading input Excel file."""
        self.image_embedder.config = {'input_file': str(self.test_input_file)}
        
        df, image_cols, b64_cols = self.image_embedder._load_input_file()
        
        self.assertIsNotNone(df)
        self.assertEqual(len(df), 3)
        
        # Check that image columns are detected
        self.assertIn('Image_1', image_cols)
        self.assertIn('Image_2', image_cols)
        self.assertIn('b64_1', b64_cols)
        self.assertIn('b64_2', b64_cols)
    
    def test_load_input_file_missing(self):
        """Test loading non-existent input file."""
        self.image_embedder.config = {'input_file': 'nonexistent.xlsx'}
        
        df, image_cols, b64_cols = self.image_embedder._load_input_file()
        
        self.assertIsNone(df)
        self.assertEqual(len(image_cols), 0)
        self.assertEqual(len(b64_cols), 0)
    
    def test_count_total_images(self):
        """Test counting total images to process."""
        test_data = {
            'Image_1': ['url1.jpg', '', 'url3.jpg'],
            'Image_2': ['', 'url2.jpg', ''],
            'b64_1': ['base64_data', '', ''],
            'b64_2': ['', 'base64_data', 'base64_data']
        }
        df = pd.DataFrame(test_data)
        image_cols = ['Image_1', 'Image_2']
        b64_cols = ['b64_1', 'b64_2']
        
        total = self.image_embedder._count_total_images(df, image_cols, b64_cols)
        
        # Image_1: 2 non-empty, Image_2: 1 non-empty, b64_1: 1, b64_2: 2
        self.assertEqual(total, 6)
    
    def test_count_total_images_empty(self):
        """Test counting images with empty dataframe."""
        test_data = {
            'Image_1': ['', '', ''],
            'b64_1': ['', '', '']
        }
        df = pd.DataFrame(test_data)
        image_cols = ['Image_1']
        b64_cols = ['b64_1']
        
        total = self.image_embedder._count_total_images(df, image_cols, b64_cols)
        self.assertEqual(total, 0)
    
    def test_validate_image_url(self):
        """Test image URL validation."""
        valid_urls = [
            'http://example.com/image.jpg',
            'https://example.com/photo.jpeg',
            'http://test.com/pic.png',
            'http://site.com/image.gif',
            'http://example.com/image.bmp',
            'http://test.com/picture.webp'
        ]
        
        invalid_urls = [
            'http://example.com/document.pdf',
            'https://example.com/video.mp4',
            'http://test.com/data.txt',
            'not-a-url',
            '',
            None
        ]
        
        for url in valid_urls:
            with self.subTest(url=url):
                self.assertTrue(validate_image_url(url))
        
        for url in invalid_urls:
            with self.subTest(url=url):
                self.assertFalse(validate_image_url(url))
    
    @patch('modules.image_embedder.requests.Session')
    def test_download_image_success(self, mock_session):
        """Test successful image download."""
        # Mock image data
        test_image_data = b'test_image_data'
        mock_response = Mock()
        mock_response.content = test_image_data
        mock_response.raise_for_status = Mock()
        
        mock_session_instance = Mock()
        mock_session_instance.get.return_value = mock_response
        mock_session.return_value = mock_session_instance
        
        self.image_embedder.configure(
            input_file=str(self.test_input_file),
            output_file=str(self.test_output_file),
            enable_cache=True
        )
        
        image_data = self.image_embedder._download_image('http://example.com/test.jpg')
        
        self.assertIsNotNone(image_data)
        self.assertEqual(image_data.getvalue(), test_image_data)
        
        # Check caching
        self.assertIn('http://example.com/test.jpg', self.image_embedder.image_cache)
    
    @patch('modules.image_embedder.requests.Session')
    def test_download_image_failure(self, mock_session):
        """Test image download failure."""
        mock_session_instance = Mock()
        mock_session_instance.get.side_effect = Exception("Network error")
        mock_session.return_value = mock_session_instance
        
        self.image_embedder.configure(
            input_file=str(self.test_input_file),
            output_file=str(self.test_output_file)
        )
        
        image_data = self.image_embedder._download_image('http://example.com/test.jpg')
        
        self.assertIsNone(image_data)
    
    def test_decode_base64_image_valid(self):
        """Test valid base64 image decoding."""
        # Create a simple test image and encode it
        img = Image.new('RGB', (50, 50), color='blue')
        img_bytes = BytesIO()
        img.save(img_bytes, format='JPEG')
        img_bytes.seek(0)
        
        encoded = base64.b64encode(img_bytes.getvalue()).decode('utf-8')
        
        # Test decoding
        decoded_data = self.image_embedder._decode_base64_image(encoded)
        
        self.assertIsNotNone(decoded_data)
        self.assertIsInstance(decoded_data, BytesIO)
    
    def test_decode_base64_image_with_data_url(self):
        """Test base64 decoding with data URL prefix."""
        # Create test image
        img = Image.new('RGB', (50, 50), color='green')
        img_bytes = BytesIO()
        img.save(img_bytes, format='JPEG')
        img_bytes.seek(0)
        
        encoded = base64.b64encode(img_bytes.getvalue()).decode('utf-8')
        data_url = f"data:image/jpeg;base64,{encoded}"
        
        # Test decoding with data URL
        decoded_data = self.image_embedder._decode_base64_image(data_url)
        
        self.assertIsNotNone(decoded_data)
        self.assertIsInstance(decoded_data, BytesIO)
    
    def test_decode_base64_image_invalid(self):
        """Test invalid base64 image decoding."""
        invalid_base64 = "invalid_base64_data!!"
        
        decoded_data = self.image_embedder._decode_base64_image(invalid_base64)
        
        self.assertIsNone(decoded_data)
    
    def test_decode_base64_image_empty(self):
        """Test empty base64 image decoding."""
        decoded_data = self.image_embedder._decode_base64_image('')
        self.assertIsNone(decoded_data)
        
        decoded_data = self.image_embedder._decode_base64_image(None)
        self.assertIsNone(decoded_data)
    
    def test_resize_image(self):
        """Test image resizing."""
        # Create a test image
        original_width, original_height = 200, 300
        img = Image.new('RGB', (original_width, original_height), color='red')
        img_bytes = BytesIO()
        img.save(img_bytes, format='JPEG')
        img_bytes.seek(0)
        
        self.image_embedder.config = {
            'max_width': 100,
            'max_height': 100,
            'image_quality': 85
        }
        
        resized_data = self.image_embedder._resize_image(img_bytes)
        
        self.assertIsNotNone(resized_data)
        
        # Check that image was resized
        with Image.open(resized_data) as resized_img:
            self.assertLessEqual(resized_img.width, 100)
            self.assertLessEqual(resized_img.height, 100)
            # Aspect ratio should be maintained
            self.assertAlmostEqual(
                original_width / original_height,
                resized_img.width / resized_img.height,
                delta=0.1
            )
    
    def test_resize_image_smaller(self):
        """Test resizing an image that's already smaller than max dimensions."""
        # Create a small test image
        img = Image.new('RGB', (50, 50), color='blue')
        img_bytes = BytesIO()
        img.save(img_bytes, format='JPEG')
        img_bytes.seek(0)
        
        self.image_embedder.config = {
            'max_width': 100,
            'max_height': 100,
            'image_quality': 85
        }
        
        resized_data = self.image_embedder._resize_image(img_bytes)
        
        self.assertIsNotNone(resized_data)
        
        # Image should remain the same size since it's already smaller
        with Image.open(resized_data) as resized_img:
            self.assertEqual(resized_img.width, 50)
            self.assertEqual(resized_img.height, 50)
    
    def test_resize_image_corrupted(self):
        """Test resizing corrupted image data."""
        corrupted_data = BytesIO(b'invalid_image_data')
        
        self.image_embedder.config = {
            'max_width': 100,
            'max_height': 100,
            'image_quality': 85
        }
        
        resized_data = self.image_embedder._resize_image(corrupted_data)
        
        self.assertIsNone(resized_data)
    
    def test_get_column_index(self):
        """Test getting column index from worksheet."""
        # Create a proper mock worksheet
        mock_ws = Mock()
        
        # Create mock cells
        mock_cell1 = Mock()
        mock_cell1.value = 'Number'
        mock_cell2 = Mock()
        mock_cell2.value = 'Image_1'
        mock_cell3 = Mock()
        mock_cell3.value = 'b64_1'
        
        # Mock the worksheet row access properly
        mock_ws.__getitem__ = Mock(return_value=[mock_cell1, mock_cell2, mock_cell3])
        
        # Test the method
        result = self.image_embedder._get_column_index(mock_ws, 'Image_1')
        self.assertEqual(result, 2)  # Should return 2 (second position)
        
        result = self.image_embedder._get_column_index(mock_ws, 'b64_1')
        self.assertEqual(result, 3)  # Should return 3 (third position)
        
        result = self.image_embedder._get_column_index(mock_ws, 'NonExistent')
        self.assertIsNone(result)  # Should return None for non-existent column
        
        
    @patch('modules.image_embedder.XLImage')
    @patch('modules.image_embedder.ImageEmbedder._resize_image')
    def test_embed_image_data_success(self, mock_resize, mock_xl_image):
        """Test successful image embedding."""
        # Mock resized image
        mock_resized_data = BytesIO(b'resized_image_data')
        mock_resize.return_value = mock_resized_data
        
        # Mock OpenPyXL image
        mock_img = Mock()
        mock_xl_image.return_value = mock_img
        
        # Mock worksheet
        mock_ws = Mock()
        
        self.image_embedder._embed_image_data(mock_ws, 2, 3, BytesIO(b'image_data'))
        
        # Verify image was created and added
        mock_xl_image.assert_called_once_with(mock_resized_data)
        mock_ws.add_image.assert_called_once_with(mock_img, 'C2')
    
    @patch('modules.image_embedder.ImageEmbedder._resize_image')
    def test_embed_image_data_resize_failure(self, mock_resize):
        """Test image embedding when resize fails."""
        mock_resize.return_value = None
        
        mock_ws = Mock()
        
        success = self.image_embedder._embed_image_data(mock_ws, 2, 3, BytesIO(b'image_data'))
        
        self.assertFalse(success)
        mock_ws.add_image.assert_not_called()
    
    @patch('modules.image_embedder.ImageEmbedder._download_image')
    @patch('modules.image_embedder.ImageEmbedder._embed_image_data')
    def test_embed_url_image_success(self, mock_embed, mock_download):
        """Test successful URL image embedding."""
        mock_download.return_value = BytesIO(b'image_data')
        mock_embed.return_value = True
        
        mock_ws = Mock()
        
        success = self.image_embedder._embed_url_image(mock_ws, 2, 3, 'http://example.com/test.jpg')
        
        self.assertTrue(success)
        mock_download.assert_called_once_with('http://example.com/test.jpg')
        mock_embed.assert_called_once()
    
    @patch('modules.image_embedder.ImageEmbedder._download_image')
    def test_embed_url_image_download_failure(self, mock_download):
        """Test URL image embedding when download fails."""
        mock_download.return_value = None
        
        mock_ws = Mock()
        
        success = self.image_embedder._embed_url_image(mock_ws, 2, 3, 'http://example.com/test.jpg')
        
        self.assertFalse(success)
        mock_download.assert_called_once_with('http://example.com/test.jpg')
    
    @patch('modules.image_embedder.ImageEmbedder._decode_base64_image')
    @patch('modules.image_embedder.ImageEmbedder._embed_image_data')
    def test_embed_base64_image_success(self, mock_embed, mock_decode):
        """Test successful base64 image embedding."""
        mock_decode.return_value = BytesIO(b'image_data')
        mock_embed.return_value = True
        
        mock_ws = Mock()
        
        success = self.image_embedder._embed_base64_image(mock_ws, 2, 3, 'base64_data')
        
        self.assertTrue(success)
        mock_decode.assert_called_once_with('base64_data')
        mock_embed.assert_called_once()
    
    @patch('modules.image_embedder.ImageEmbedder._decode_base64_image')
    def test_embed_base64_image_decode_failure(self, mock_decode):
        """Test base64 image embedding when decode fails."""
        mock_decode.return_value = None
        
        mock_ws = Mock()
        
        success = self.image_embedder._embed_base64_image(mock_ws, 2, 3, 'base64_data')
        
        self.assertFalse(success)
        mock_decode.assert_called_once_with('base64_data')
    
    def test_prepare_workbook_new(self):
        """Test preparing a new workbook."""
        self.image_embedder.config = {'output_file': str(self.test_output_file)}
        
        wb, ws = self.image_embedder._prepare_workbook()
        
        self.assertIsNotNone(wb)
        self.assertIsNotNone(ws)
        self.assertIsInstance(wb, openpyxl.Workbook)
    
    def test_prepare_workbook_existing(self):
        """Test preparing an existing workbook."""
        # First create a workbook
        wb = openpyxl.Workbook()
        wb.save(self.test_output_file)
        
        self.image_embedder.config = {'output_file': str(self.test_output_file)}
        
        wb, ws = self.image_embedder._prepare_workbook()
        
        self.assertIsNotNone(wb)
        self.assertIsNotNone(ws)
    
    def test_configure_excel_layout(self):
        """Test Excel layout configuration."""
        # Create a test workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add some data
        ws['A1'] = 'Number'
        ws['B1'] = 'Image_1'
        ws['C1'] = 'b64_1'
        for i in range(2, 5):  # Add some rows
            ws[f'A{i}'] = f'Number{i}'
        
        self.image_embedder.config = {
            'row_height': 80,
            'column_width': 15
        }
        
        image_cols = ['Image_1']
        b64_cols = ['b64_1']
        
        self.image_embedder._configure_excel_layout(ws, image_cols, b64_cols)
        
        # Check row heights
        for i in range(2, 5):
            self.assertEqual(ws.row_dimensions[i].height, 80)
        
        # Check column widths (approximate due to Excel units)
        self.assertIsNotNone(ws.column_dimensions['B'].width)
        self.assertIsNotNone(ws.column_dimensions['C'].width)
    
    def test_save_workbook(self):
        """Test saving workbook."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Test Data'
        
        self.image_embedder.config = {'output_file': str(self.test_output_file)}
        
        success = self.image_embedder._save_workbook(wb)
        
        self.assertTrue(success)
        self.assertTrue(self.test_output_file.exists())
    
    def test_save_workbook_permission_error(self):
        """Test saving workbook with permission error."""
        wb = openpyxl.Workbook()
        
        # Create a read-only directory to cause permission error
        with tempfile.TemporaryDirectory() as temp_dir:
            read_only_file = Path(temp_dir) / 'readonly.xlsx'
            read_only_file.touch()
            read_only_file.chmod(0o444)  # Read-only
            
            self.image_embedder.config = {'output_file': str(read_only_file)}
            
            success = self.image_embedder._save_workbook(wb)
            
            self.assertFalse(success)
    
    def test_get_processing_stats(self):
        """Test getting processing statistics."""
        self.image_embedder.processed_count = 15
        self.image_embedder.error_count = 3
        self.image_embedder.is_running = True
        
        stats = self.image_embedder.get_processing_stats()
        
        self.assertEqual(stats['processed_count'], 15)
        self.assertEqual(stats['error_count'], 3)
        self.assertTrue(stats['is_running'])
    
    def test_stop_method(self):
        """Test stopping the image embedder process."""
        self.image_embedder.is_running = True
        self.image_embedder.stop()
        
        self.assertFalse(self.image_embedder.is_running)


class TestImageEmbedderIntegration(unittest.TestCase):
    """Integration tests for ImageEmbedder."""
    
    def setUp(self):
        """Set up integration test fixtures."""
        self.image_embedder = ImageEmbedder()
        self.test_input_file = Path(tempfile.mktemp(suffix=".xlsx"))
        self.test_output_file = Path(tempfile.mktemp(suffix=".xlsx"))
        
        # Create test data with base64 images (no network required)
        self._create_test_excel_with_base64()
    
    def tearDown(self):
        """Clean up integration test files."""
        for file_path in [self.test_input_file, self.test_output_file]:
            if file_path.exists():
                file_path.unlink()
    
    def _create_test_excel_with_base64(self):
        """Create test Excel file with base64 images."""
        # Create a simple test image and encode it
        img = Image.new('RGB', (50, 50), color='red')
        img_bytes = BytesIO()
        img.save(img_bytes, format='JPEG')
        img_bytes.seek(0)
        encoded = base64.b64encode(img_bytes.getvalue()).decode('utf-8')
        
        test_data = {
            'Number': ['923001234567', '923001234568'],
            'Name': ['User One', 'User Two'],
            'b64_1': [encoded, ''],
            'b64_2': ['', encoded]
        }
        df = pd.DataFrame(test_data)
        df.to_excel(self.test_input_file, index=False)
    
    def test_complete_processing_flow_base64(self):
        """Test complete processing flow with base64 images."""
        # Configure callbacks
        mock_log = Mock()
        mock_status = Mock()
        mock_progress = Mock()
        mock_stop = Mock(return_value=False)
        
        self.image_embedder.configure(
            input_file=str(self.test_input_file),
            output_file=str(self.test_output_file),
            max_width=40,
            max_height=40,
            log_callback=mock_log,
            status_callback=mock_status,
            progress_callback=mock_progress,
            stop_callback=mock_stop,
            enable_cache=False  # Disable cache to avoid session creation
        )
        
        # Run processing
        success = self.image_embedder.run()
        
        # Verify results
        self.assertTrue(success)
        self.assertTrue(self.test_output_file.exists())
        
        # Verify callbacks were called
        self.assertTrue(mock_log.called)
        self.assertTrue(mock_status.called)
        self.assertTrue(mock_progress.called)
        
        # Verify output file can be opened and has images
        wb = openpyxl.load_workbook(self.test_output_file)
        ws = wb.active
        
        # Check that images were embedded (they should be in the worksheet)
        self.assertGreater(len(ws._images), 0)
        
        wb.close()
        
    def test_processing_stop_signal(self):
        """Test processing with stop signal."""
        mock_stop = Mock()
        # Return False first few times, then True to simulate stop signal
        # The method calls should_stop() multiple times during processing
        mock_stop.side_effect = [False, False, False, True]  # More calls to avoid StopIteration

        self.image_embedder.configure(
            input_file=str(self.test_input_file),
            output_file=str(self.test_output_file),
            stop_callback=mock_stop,
            enable_cache=False
        )

        success = self.image_embedder.run()

        # Should complete but may not process all images due to stop signal
        self.assertTrue(success)  # Should return True as it stopped gracefully

class TestUtilityFunctions(unittest.TestCase):
    """Test utility functions."""
    
    def test_embed_images_in_excel(self):
        """Test convenience function for image embedding."""
        # Create test input file
        test_input = Path(tempfile.mktemp(suffix=".xlsx"))
        test_output = Path(tempfile.mktemp(suffix=".xlsx"))
        
        # Create simple test data
        test_data = {
            'Number': ['923001234567'],
            'b64_1': ['']
        }
        df = pd.DataFrame(test_data)
        df.to_excel(test_input, index=False)
        
        # Test the function
        success = embed_images_in_excel(
            input_file=str(test_input),
            output_file=str(test_output),
            max_width=50,
            max_height=50
        )
        
        self.assertTrue(success)
        self.assertTrue(test_output.exists())
        
        # Clean up
        for file_path in [test_input, test_output]:
            if file_path.exists():
                file_path.unlink()
    
    def test_embed_images_in_excel_auto_output(self):
        """Test convenience function with auto-generated output filename."""
        # Create test input file
        test_input = Path(tempfile.mktemp(suffix=".xlsx"))
        
        # Create simple test data
        test_data = {'Number': ['923001234567']}
        df = pd.DataFrame(test_data)
        df.to_excel(test_input, index=False)
        
        # Test without output file (should auto-generate)
        success = embed_images_in_excel(
            input_file=str(test_input),
            max_width=50,
            max_height=50
        )
        
        self.assertTrue(success)
        
        # Check that output file was created with expected name
        expected_output = test_input.parent / f"{test_input.stem}_with_images{test_input.suffix}"
        self.assertTrue(expected_output.exists())
        
        # Clean up
        for file_path in [test_input, expected_output]:
            if file_path.exists():
                file_path.unlink()


class TestErrorScenarios(unittest.TestCase):
    """Test error scenarios and edge cases."""
    
    def test_invalid_input_file(self):
        """Test handling of invalid input file."""
        embedder = ImageEmbedder()
        
        embedder.configure(
            input_file='nonexistent_file.xlsx',
            output_file='output.xlsx'
        )
        
        success = embedder.run()
        self.assertFalse(success)
    
    def test_empty_dataframe(self):
        """Test handling of empty input dataframe."""
        embedder = ImageEmbedder()
        
        # Create empty test file
        test_file = Path(tempfile.mktemp(suffix=".xlsx"))
        df = pd.DataFrame()
        df.to_excel(test_file, index=False)
        
        embedder.configure(
            input_file=str(test_file),
            output_file='output.xlsx'
        )
        
        # Should complete successfully but process 0 images
        success = embedder.run()
        self.assertTrue(success)
        
        # Clean up
        if test_file.exists():
            test_file.unlink()
    
    def test_no_image_columns(self):
        """Test processing file with no image columns."""
        embedder = ImageEmbedder()
        
        # Create test file without image columns
        test_file = Path(tempfile.mktemp(suffix=".xlsx"))
        test_data = {
            'Number': ['923001234567'],
            'Name': ['User One']
        }
        df = pd.DataFrame(test_data)
        df.to_excel(test_file, index=False)
        
        embedder.configure(
            input_file=str(test_file),
            output_file='output.xlsx'
        )
        
        # Should complete successfully but process 0 images
        success = embedder.run()
        self.assertTrue(success)
        
        # Clean up
        if test_file.exists():
            test_file.unlink()


if __name__ == '__main__':
    # Create test data directory
    test_data_dir = Path(__file__).parent / "test_data"
    test_data_dir.mkdir(exist_ok=True)
    
    # Run tests
    unittest.main(verbosity=2)